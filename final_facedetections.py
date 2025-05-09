import cv2
import sqlite3
import os
import numpy as np
import face_recognition
from datetime import datetime, time as dt_time, timedelta
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import bcrypt
from deepface import DeepFace
import pyttsx3
import threading
import time
import schedule
from tkinter import font as tkfont
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import calendar

# ========== KONFIGURASI ==========
FACE_DETECTION_MODEL = "hog"  # "hog" untuk CPU, "cnn" untuk GPU
FACE_MATCHING_TOLERANCE = 0.6
TARGET_FRAME_WIDTH = 640
UPDATE_INTERVAL = 300  # 5 menit dalam detik
ROLE_OPTIONS = ['Guru', 'Siswa', 'Staf', 'Admin']
LEVEL_OPTIONS = ['7', '8', '9', '10']
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"  # Password akan di-hash

# Waktu presensi
MORNING_ENTRY_START = dt_time(5, 0)   # 05:00 WIB mulai presensi masuk
MORNING_ENTRY_END = dt_time(9, 0)     # 09:00 WIB batas akhir presensi masuk
AFTERNOON_EXIT_START = dt_time(13, 0) # 13:00 WIB mulai presensi pulang
AFTERNOON_EXIT_END = dt_time(16, 0)   # 16:00 WIB batas akhir presensi pulang
GRACE_PERIOD = 15  # toleransi 15 menit

# ========== INISIALISASI DIREKTORI ==========
def init_directories():
    os.makedirs('img', exist_ok=True) # folder tempat menyimpan gambar wajah pengguna.
    os.makedirs('known_faces', exist_ok=True) # folder tempat menyimpan gambar wajah yang sudah terdaftar.
    os.makedirs('screenshots', exist_ok=True) # folder tempat menyimpan screenshot.
    os.makedirs('attendance_screenshots', exist_ok=True) # folder tempat menyimpan screenshot kehadiran.
    os.makedirs('attendance_records', exist_ok=True) # folder tempat menyimpan catatan kehadiran.
    os.makedirs('excel_reports', exist_ok=True) # folder tempat menyimpan laporan excel.
    
    # Buat direktori untuk screenshot berdasarkan peran
    roles = ['siswa', 'guru', 'staf', 'admin']
    for role in roles:
        role_path = os.path.join('screenshots', role)
        os.makedirs(role_path, exist_ok=True)
        
        # Untuk siswa, buat subdirektori berdasarkan kelas
        if role == 'siswa':
            for level in ['7', '8', '9', '10']:
                level_path = os.path.join(role_path, f'kelas_{level}')
                os.makedirs(level_path, exist_ok=True)

init_directories()

# ========== STYLE GUI ==========
class AppStyle:
    """Kelas untuk mengatur style (warna, font) aplikasi dan menerapkannya ke root window."""
    def __init__(self):
        # Inisialisasi warna dan font
        self.bg_color = "#f0f0f0"
        self.primary_color = "#4a6fa5"
        self.secondary_color = "#166088"
        self.accent_color = "#4fc3f7"
        self.text_color = "#333333"
        self.error_color = "#d32f2f"
        self.success_color = "#388e3c"
        
        self.title_font = ("Helvetica", 16, "bold")
        self.subtitle_font = ("Helvetica", 12)
        self.normal_font = ("Helvetica", 10)
        
    def apply_style(self, root):
        """Terapkan style ke root window dan widget utama."""
        # Set root window properties
        root.configure(bg=self.bg_color)
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure styles
        style.configure('TFrame', background=self.bg_color) # Set background color for frames
        style.configure('TLabel', background=self.bg_color, foreground=self.text_color, font=self.normal_font) # Set background color for labels
        style.configure('TButton', font=self.normal_font, background=self.primary_color, foreground='white') # Set background color for buttons
        style.configure('TEntry', font=self.normal_font) # Set background color for entry fields
        style.configure('TCombobox', font=self.normal_font) # Set background color for comboboxes
        
        style.map('TButton',
                background=[('active', self.secondary_color), ('pressed', self.accent_color)],
                foreground=[('active', 'white')]) # Set button hover and pressed colors
        
        style.configure('Accent.TButton', background=self.accent_color) # Set accent button color
        style.configure('Error.TLabel', foreground=self.error_color) # Set error label color
        style.configure('Success.TLabel', foreground=self.success_color) # # Set success label color

# ========== DATABASE ==========
class DailyReport:
    """Kelas untuk membuat laporan harian dalam format Excel."""
    def __init__(self):
        # Inisialisasi direktori laporan
        self.report_dir = 'daily_reports'
        os.makedirs(self.report_dir, exist_ok=True)
        
    def get_day_name(self, date):
        """Mendapatkan nama hari dalam bahasa Indonesia dari objek date."""
        day_names = {
            0: 'Senin',
            1: 'Selasa',
            2: 'Rabu',
            3: 'Kamis',
            4: 'Jumat',
            5: 'Sabtu',
            6: 'Minggu'
        }
        return day_names[date.weekday()]
        
    def create_daily_report(self, date=None):
        """Membuat file laporan harian Excel untuk tanggal tertentu."""
        if date is None:
            date = datetime.now()
            
        # Format nama file: Senin_2024-03-18.xlsx
        day_name = self.get_day_name(date)
        filename = f"{day_name}_{date.strftime('%Y-%m-%d')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        
        # Buat workbook baru
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Kehadiran"
        
        # Header style
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")
        
        # Set header
        ws['A1'] = f"Laporan Kehadiran {day_name}, {date.strftime('%d %B %Y')}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Set kolom header
        headers = ['Nama', 'Level', 'Waktu Kehadiran', 'Emosi Saat Datang', 
                  'Waktu Kepulangan', 'Emosi Saat Pulang', 'Durasi di Sekolah']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Set lebar kolom
        column_widths = {
            'A': 30,  # Nama
            'B': 15,  # Level
            'C': 20,  # Waktu Kehadiran
            'D': 20,  # Emosi Saat Datang
            'E': 20,  # Waktu Kepulangan
            'F': 20,  # Emosi Saat Pulang
            'G': 20   # Durasi
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        return filepath, wb, ws

class FaceDatabase:
    """Kelas untuk mengelola database wajah, admin, dan presensi."""
    def __init__(self):
        # Inisialisasi koneksi ke database
        self.conn = sqlite3.connect('faces.db')
        self.admin_conn = sqlite3.connect('admin.db')
        self.attendance_conn = sqlite3.connect('attendance.db')
        self.lock = threading.Lock()
        self.known_face_encodings = []
        self.known_face_names = []
        self.known_face_roles = []
        self.known_face_levels = []
        self.last_update = 0
        self.daily_report = DailyReport()
        self.init_db()
        self.load_known_faces()

    def init_db(self):
        """Membuat tabel-tabel database jika belum ada."""
        with self.lock:
            cursor = self.conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS faces (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    role TEXT,
                    level TEXT DEFAULT '',
                    image_path TEXT,
                    encoding BLOB,
                    timestamp TEXT,
                    emotion TEXT
                )
            ''')
            self.conn.commit()

            admin_cursor = self.admin_conn.cursor()
            admin_cursor.execute('''
                CREATE TABLE IF NOT EXISTS admin (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL
                )
            ''')
            
            # Hash password admin
            hashed_password = bcrypt.hashpw(ADMIN_PASSWORD.encode('utf-8'), bcrypt.gensalt())
            admin_cursor.execute("INSERT OR IGNORE INTO admin (username, password) VALUES (?, ?)", 
                               (ADMIN_USERNAME, hashed_password))
            self.admin_conn.commit()
            
            # Tabel presensi dengan kolom tambahan
            attendance_cursor = self.attendance_conn.cursor()
            attendance_cursor.execute('''
                CREATE TABLE IF NOT EXISTS attendance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    role TEXT NOT NULL,
                    level TEXT,
                    date TEXT NOT NULL,
                    entry_time TEXT,
                    exit_time TEXT,
                    entry_emotion TEXT,
                    exit_emotion TEXT,
                    duration TEXT,
                    status TEXT,
                    keterangan TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            self.attendance_conn.commit()

    def load_known_faces(self):
        """Memuat data wajah yang dikenal dari database ke memori."""
        with self.lock:
            self.known_face_encodings = []
            self.known_face_names = []
            self.known_face_roles = []
            self.known_face_levels = []

            cursor = self.conn.cursor()
            # Periksa apakah kolom level ada
            cursor.execute("PRAGMA table_info(faces)")
            columns = [column[1] for column in cursor.fetchall()]
            has_level = 'level' in columns

            if has_level:
                cursor.execute("SELECT name, role, level, encoding FROM faces")
                for name, role, level, encoding_blob in cursor.fetchall():
                    encoding = np.frombuffer(encoding_blob, dtype=np.float64)
                    self.known_face_encodings.append(encoding)
                    self.known_face_names.append(name)
                    self.known_face_roles.append(role)
                    self.known_face_levels.append(level)
            else:
                # Jika kolom level belum ada, tambahkan kolom
                cursor.execute("ALTER TABLE faces ADD COLUMN level TEXT DEFAULT ''")
                self.conn.commit()
                cursor.execute("SELECT name, role, encoding FROM faces")
                for name, role, encoding_blob in cursor.fetchall():
                    encoding = np.frombuffer(encoding_blob, dtype=np.float64)
                    self.known_face_encodings.append(encoding)
                    self.known_face_names.append(name)
                    self.known_face_roles.append(role)
                    self.known_face_levels.append('')  # Default empty level
            
            self.last_update = time.time()
            print(f"Memuat {len(self.known_face_names)} wajah yang dikenal")

    def add_face(self, name, role, level, face_image, face_encoding, emotion):
        """Menambahkan wajah baru ke database dan memori."""
        with self.lock:
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            img_filename = f"known_faces/{name}_{timestamp}.jpg"
            cv2.imwrite(img_filename, face_image)
            
            encoding_blob = face_encoding.tobytes()
            
            cursor = self.conn.cursor()
            cursor.execute('''
                INSERT INTO faces (name, role, level, image_path, encoding, timestamp, emotion)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (name, role, level, img_filename, encoding_blob, timestamp, emotion))
            self.conn.commit()
            
            # Update data di memori
            self.known_face_encodings.append(face_encoding)
            self.known_face_names.append(name)
            self.known_face_roles.append(role)
            self.known_face_levels.append(level)
            print(f"Wajah {name} berhasil didaftarkan sebagai {role} level {level}")

    def verify_admin(self, username, password):
        """Memverifikasi login admin dengan username dan password."""
        cursor = self.admin_conn.cursor()
        cursor.execute("SELECT password FROM admin WHERE username = ?", (username,))
        result = cursor.fetchone()
        
        if result:
            return bcrypt.checkpw(password.encode('utf-8'), result[0])
        return False
    
    def record_attendance(self, name, role, level, is_entry=True, emotion=None, keterangan=None):
        """Mencatat kehadiran atau kepulangan ke database."""
        today = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H:%M:%S')
        
        with self.lock:
            cursor = self.attendance_conn.cursor()
            
            if is_entry:
                # Rekam presensi masuk sebagai record baru
                cursor.execute('''
                    INSERT INTO attendance (name, role, level, date, entry_time, entry_emotion, status, keterangan)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, role, level, today, current_time, emotion, 'Hadir', keterangan))
                self.attendance_conn.commit()
                print(f"Presensi masuk {name} dicatat pada {current_time}")
            else:
                # Cek record terakhir untuk orang yang sama
                cursor.execute('''
                    SELECT id, entry_time FROM attendance 
                    WHERE name = ? AND date = ? 
                    ORDER BY entry_time DESC LIMIT 1
                ''', (name, today))
                
                result = cursor.fetchone()
                if result:
                    record_id, entry_time_str = result
                    if entry_time_str:  # Pastikan entry_time tidak None
                        try:
                            entry_time = datetime.strptime(f"{today} {entry_time_str}", "%Y-%m-%d %H:%M:%S")
                            exit_time = datetime.strptime(f"{today} {current_time}", "%Y-%m-%d %H:%M:%S")
                            duration = exit_time - entry_time
                            
                            # Format durasi menjadi HH:MM:SS
                            hours, remainder = divmod(duration.seconds, 3600)
                            minutes, seconds = divmod(remainder, 60)
                            duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                            
                            # Update record dengan waktu pulang
                            cursor.execute('''
                                UPDATE attendance 
                                SET exit_time = ?, exit_emotion = ?, duration = ?, status = 'Pulang', keterangan = ?
                                WHERE id = ?
                            ''', (current_time, emotion, duration_str, keterangan, record_id))
                            self.attendance_conn.commit()
                            print(f"Presensi pulang {name} dicatat pada {current_time}")
                        except ValueError as e:
                            print(f"Error format waktu: {e}")
                            # Jika ada error format waktu, buat record baru
                            cursor.execute('''
                                INSERT INTO attendance (name, role, level, date, exit_time, exit_emotion, status, keterangan)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (name, role, level, today, current_time, emotion, 'Pulang (Error Format)', keterangan))
                            self.attendance_conn.commit()
                    else:
                        # Jika entry_time None, buat record baru
                        cursor.execute('''
                            INSERT INTO attendance (name, role, level, date, exit_time, exit_emotion, status, keterangan)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (name, role, level, today, current_time, emotion, 'Pulang Tanpa Masuk', keterangan))
                        self.attendance_conn.commit()
                else:
                    # Jika tidak ada record kedatangan, buat record baru untuk kepulangan
                    cursor.execute('''
                        INSERT INTO attendance (name, role, level, date, exit_time, exit_emotion, status, keterangan)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (name, role, level, today, current_time, emotion, 'Pulang Tanpa Masuk', keterangan))
                    self.attendance_conn.commit()
                    print(f"Presensi pulang {name} (tanpa masuk) dicatat pada {current_time}")

    def export_to_excel(self, date=None):
        """Ekspor data presensi ke file Excel untuk tanggal tertentu."""
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        
        with self.lock:
            cursor = self.attendance_conn.cursor()
            cursor.execute('''
                SELECT name, role, level, date, entry_time, exit_time, duration, entry_emotion, exit_emotion, status
                FROM attendance
                WHERE date = ?
                ORDER BY name
            ''', (date,))
            
            data = cursor.fetchall()
            
            if not data:
                return False
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=[
                'Nama', 'Role', 'Level', 'Tanggal', 
                'Jam Masuk', 'Jam Pulang', 'Durasi', 
                'Emosi Saat Datang', 'Emosi Saat Pulang', 'Status'
            ])
            
            # Create Excel file
            filename = f"excel_reports/attendance_report_{date}.xlsx"
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Attendance')
            
            # Get workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Style header
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filename)
            return filename

    def get_attendance_data(self, date):
        """Mengambil data presensi dari database untuk tanggal tertentu."""
        with self.lock:
            cursor = self.attendance_conn.cursor()
            cursor.execute('''
                SELECT name, role, level, date, entry_time, exit_time, status
                FROM attendance
                WHERE date = ?
                ORDER BY name
            ''', (date,))
            
            columns = ['name', 'role', 'level', 'date', 'entry_time', 'exit_time', 'status']
            data = []
            
            for row in cursor.fetchall():
                data.append(dict(zip(columns, row)))
                
            return data

    def generate_daily_report(self, date=None):
        """Membuat laporan harian Excel untuk tanggal tertentu."""
        if date is None:
            date = datetime.now()
            
        filepath, wb, ws = self.daily_report.create_daily_report(date)
        
        # Ambil data dari database
        with self.lock:
            cursor = self.attendance_conn.cursor()
            cursor.execute('''
                SELECT name, level, entry_time, entry_emotion, 
                       exit_time, exit_emotion, duration
                FROM attendance
                WHERE date = ?
                ORDER BY name, entry_time
            ''', (date.strftime('%Y-%m-%d'),))
            
            rows = cursor.fetchall()
            
            # Tulis data ke Excel
            for row_idx, row in enumerate(rows, 3):  # Mulai dari baris 3
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Handle nilai None
                    if value is None:
                        cell.value = '-'
                    else:
                        cell.value = value
                    cell.alignment = Alignment(horizontal='center')
        
        # Simpan file
        wb.save(filepath)
        return filepath

    def update_admin_credentials(self, new_username, new_password):
        import bcrypt
        cursor = self.admin_conn.cursor()
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
        try:
            cursor.execute("UPDATE admin SET username = ?, password = ? WHERE id = 1", (new_username, hashed_password))
            self.admin_conn.commit()
            return True
        except Exception as e:
            print(f"Error updating admin credentials: {e}")
            return False

# ========== ADMIN LOGIN GUI ==========
class AdminLoginWindow:
    def __init__(self, parent, face_db):
        self.parent = parent
        self.face_db = face_db
        self.style = AppStyle()
        self.authenticated = False
        
        self.root = tk.Toplevel(parent)
        self.root.title("Admin Login - Praxis High School")
        self.root.geometry("450x500")
        self.root.resizable(False, False)
        self.style.apply_style(self.root)
        
        # Create a canvas for the logo and school name
        self.canvas = tk.Canvas(self.root, width=400, height=150, bg=self.style.bg_color, highlightthickness=0)
        self.canvas.pack(pady=(20, 10))
        
        # Draw school logo (placeholder - replace with actual logo)
        self.draw_logo()
        
        self.create_widgets()
        
    def draw_logo(self):
        try:
            # Try to load actual logo
            logo_img = Image.open("praxcis-removebg-preview.png")
            logo_img = logo_img.resize((100, 100), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_img)
            
            # Create circular mask for logo
            self.canvas.create_image(200, 50, image=self.logo_photo)
        except:
            # Fallback to simple circle if logo not found
            self.canvas.create_oval(150, 0, 250, 100, fill=self.style.primary_color, outline="")
        
        # Add school name
        self.canvas.create_text(200, 120, text="Praxis High School", 
                              font=("Helvetica", 14, "bold"), 
                              fill=self.style.primary_color)
        
    def create_widgets(self):
        # Main container
        container = ttk.Frame(self.root, style='TFrame')
        container.pack(fill=tk.BOTH, expand=True, padx=30, pady=10)
        
        # Login title
        ttk.Label(container, text="ADMINISTRATOR LOGIN", 
                 font=self.style.title_font).pack(pady=(10, 20))
        
        # Form frame
        form_frame = ttk.Frame(container, style='TFrame')
        form_frame.pack(fill=tk.X, pady=10)
        
        # Username
        ttk.Label(form_frame, text="Username:", font=self.style.subtitle_font).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.username_entry = ttk.Entry(form_frame, font=self.style.normal_font)
        self.username_entry.grid(row=0, column=1, sticky=tk.EW, pady=5, padx=5)
        
        # Password
        ttk.Label(form_frame, text="Password:", font=self.style.subtitle_font).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.password_entry = ttk.Entry(form_frame, show="•", font=self.style.normal_font)
        self.password_entry.grid(row=1, column=1, sticky=tk.EW, pady=5, padx=5)
        
        # Error message
        self.error_label = ttk.Label(form_frame, text="", style='Error.TLabel')
        self.error_label.grid(row=2, column=0, columnspan=2, pady=5)
        
        # Button frame
        button_frame = ttk.Frame(container, style='TFrame')
        button_frame.pack(fill=tk.X, pady=20)
        
        # Login button
        login_btn = ttk.Button(button_frame, text="LOGIN", command=self.attempt_login, 
                             style='Accent.TButton')
        login_btn.pack(fill=tk.X, pady=5, ipady=5)
        
        # Footer
        footer_frame = ttk.Frame(self.root, style='TFrame')
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
        
        ttk.Label(footer_frame, text="© 2025 Praxis High School Attendance System", 
                 font=("Helvetica", 8), style='TLabel').pack()
        
        # Bind Enter key
        self.root.bind('<Return>', lambda e: self.attempt_login())
        self.username_entry.focus()
        
    def attempt_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        if not username or not password:
            self.error_label.config(text="Username dan password harus diisi!")
            return
            
        if self.face_db.verify_admin(username, password):
            self.authenticated = True
            self.root.destroy()
        else:
            self.error_label.config(text="Username atau password salah!")

# ========== ATTENDANCE SYSTEM ==========
class AttendanceSystem:
    """Kelas utama untuk logika presensi, suara, dan screenshot."""
    def __init__(self, face_db):
        self.face_db = face_db
        self.voice_engine = pyttsx3.init()
        self.voice_engine.setProperty('rate', 150)
        self.recognized_faces = {}
        self.screenshot_taken = set()  # Untuk melacak screenshot yang sudah diambil
        self.schedule_jobs()
        self.last_confirmation_time = {}  # Untuk mencegah konfirmasi berulang
        self.last_exit_time = {}  # Tambah untuk pembatasan presensi pulang
        
    def schedule_jobs(self):
        """Menjadwalkan tugas harian (reset data, ekspor, laporan)."""
        # Jadwalkan pembersihan data harian
        schedule.every().day.at("00:00").do(self.clear_daily_data)
        # Hapus penjadwalan ekspor dan laporan harian otomatis
        # schedule.every().day.at("23:59").do(self.export_daily_report)
        # schedule.every().day.at("05:00").do(self.generate_daily_report)
        
    def clear_daily_data(self):
        """Membersihkan data harian (cache presensi, screenshot, dsb)."""
        self.recognized_faces.clear()
        self.screenshot_taken.clear()  # Bersihkan juga daftar screenshot
        self.last_confirmation_time.clear()  # Bersihkan waktu konfirmasi terakhir
        print("Data harian telah dibersihkan")
    
    def export_daily_report(self):
        """Ekspor laporan presensi harian ke Excel (untuk kemarin)."""
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        filename = self.face_db.export_to_excel(yesterday)
        if filename:
            print(f"Laporan harian berhasil diekspor ke: {filename}")
        else:
            print("Tidak ada data presensi untuk diekspor")
        
    def generate_daily_report(self):
        """Membuat laporan harian Excel untuk tanggal tertentu."""
        """Membuat laporan untuk hari sebelumnya"""
        yesterday = datetime.now() - timedelta(days=1)
        filepath = self.face_db.generate_daily_report(yesterday)
        print(f"Laporan harian berhasil dibuat: {filepath}")
        
    def check_attendance_time(self, frame, name, role, level, parent_window=None):
        """Cek dan proses presensi masuk sesuai waktu dan konfirmasi."""
        current_time = datetime.now().time()
        today = datetime.now().strftime('%Y-%m-%d')
        emotion = detect_emotion(frame)
        
        # Cek apakah sudah waktunya untuk konfirmasi baru
        key = f"{name}_{today}"
        last_time = self.last_confirmation_time.get(key, datetime.min)
        time_diff = datetime.now() - last_time
        
        # Jika belum 5 menit sejak konfirmasi terakhir, skip
        if time_diff.total_seconds() < 300:  # 5 menit
            return
        
        # Cek waktu presensi masuk (05:00-09:00)
        if (self.is_time_between(current_time, MORNING_ENTRY_START, MORNING_ENTRY_END)):
            # Tampilkan window konfirmasi
            confirmation_window = ConfirmationWindow(
                parent_window, name, role, level, emotion, is_entry=True
            )
            parent_window.wait_window(confirmation_window.root)
            
            if confirmation_window.confirmed:
                # Catat kedatangan
                self.face_db.record_attendance(
                    name, role, level, 
                    is_entry=True, 
                    emotion=emotion,
                    keterangan=confirmation_window.keterangan
                )
                
                # Update waktu konfirmasi terakhir
                self.last_confirmation_time[key] = datetime.now()
                
                # Ambil screenshot
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"attendance_screenshots/ENTRY_{name}_{timestamp}.jpg"
                cv2.imwrite(filename, frame)
                
                # Efek suara kedatangan
                self.voice_engine.say(f"Welcome {name}, Your current emotion is {emotion}. Good luck with your studies.")
                self.voice_engine.runAndWait()
    
    def is_time_between(self, check_time, start_time, end_time):
        """Cek apakah waktu saat ini berada di antara rentang waktu tertentu."""
        if start_time <= end_time:
            return start_time <= check_time <= end_time
        else:  # Waktu melewati tengah malam
            return start_time <= check_time or check_time <= end_time

    def take_screenshot(self, frame, name, role, level):
        """Ambil screenshot presensi sesuai peran dan kelas."""
        today = datetime.now().strftime('%Y-%m-%d')
        key = f"{name}_{today}"
        
        if key not in self.screenshot_taken:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Tentukan path berdasarkan peran
            role = role.lower()
            if role == 'siswa':
                # Untuk siswa, simpan di folder kelas yang sesuai
                filename = f"screenshots/{role}/kelas_{level}/{name}_{timestamp}.jpg"
            else:
                # Untuk peran lain (guru, staf, admin)
                filename = f"screenshots/{role}/{name}_{timestamp}.jpg"
            
            cv2.imwrite(filename, frame)
            self.screenshot_taken.add(key)
            print(f"Screenshot diambil untuk {name} ({role})")
    
    def export_screenshot_report(self, date=None, role=None, level=None):
        """Ekspor laporan screenshot ke Excel."""
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
            
        # Buat DataFrame untuk menyimpan data screenshot
        data = []
        
        # Tentukan path berdasarkan peran
        if role:
            role = role.lower()
            if role == 'siswa' and level:
                screenshot_dir = f"screenshots/{role}/kelas_{level}"
            else:
                screenshot_dir = f"screenshots/{role}"
        else:
            screenshot_dir = "screenshots"
            
        # Cari semua file screenshot
        for root, dirs, files in os.walk(screenshot_dir):
            for file in files:
                if file.endswith('.jpg'):
                    file_path = os.path.join(root, file)
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    
                    # Ekstrak informasi dari nama file
                    name = file.split('_')[0]
                    role_from_path = root.split(os.sep)[1]
                    level_from_path = root.split(os.sep)[2] if len(root.split(os.sep)) > 2 else None
                    
                    data.append({
                        'Nama': name,
                        'Peran': role_from_path,
                        'Kelas': level_from_path,
                        'Waktu': file_time.strftime('%H:%M:%S'),
                        'Tanggal': file_time.strftime('%Y-%m-%d'),
                        'Path': file_path
                    })
        
        if not data:
            return False
            
        # Buat DataFrame dan ekspor ke Excel
        df = pd.DataFrame(data)
        filename = f"excel_reports/screenshot_report_{date}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Screenshots')
            
            # Styling
            workbook = writer.book
            worksheet = writer.sheets['Screenshots']
            
            # Header style
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
                
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        return filename

# ========== MAIN APPLICATION ==========
class FaceRecognitionApp:
    """Kelas utama aplikasi GUI presensi berbasis face recognition."""
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("School Attendance System (MTS-MA)")
        self.is_fullscreen = False
        self.root.geometry("1400x800")
        self.style = AppStyle()  # Pastikan style diinisialisasi
        # Inisialisasi database dan sistem presensi lebih awal
        self.face_db = FaceDatabase()
        self.face_db.init_db()
        self.face_db.load_known_faces()
        self.attendance_system = AttendanceSystem(self.face_db)
        self.admin_mode = False
        self.admin_window = None
        self.current_frame = None
        self.current_face_location = None
        self.current_face_encoding = None
        self.last_manual_confirmation = {}
        # Frame judul dan waktu
        self.header_frame = ttk.Frame(self.root)
        self.header_frame.pack(side=tk.TOP, fill=tk.X)
        self.title_label = ttk.Label(self.header_frame, text="School Attendance System (MTS-MA)",
                                    font=("Helvetica", 28, "bold"), foreground="#1660d0", anchor="center")
        self.title_label.pack(pady=(10, 0))
        self.time_label = ttk.Label(self.header_frame, text="", font=("Helvetica", 20, "bold"), foreground="#333399")
        self.time_label.pack(pady=(0, 10))
        self.fullscreen_btn = ttk.Button(self.header_frame, text="Fullscreen", command=self.toggle_fullscreen)
        self.fullscreen_btn.place(relx=1.0, x=-20, y=10, anchor="ne")
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        self.shortcut_frame = ttk.Frame(self.main_frame, width=300)
        self.shortcut_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)
        self.camera_frame = ttk.Frame(self.main_frame)
        self.camera_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.camera_label = ttk.Label(self.camera_frame)
        self.camera_label.pack(fill=tk.BOTH, expand=True)
        self.create_shortcut_panel()
        self.root.bind('<Key>', self.handle_key_press)
        self.cap = cv2.VideoCapture(0)
        self.update_camera()
        self.update_time_label()  # update waktu
        self.scheduler_thread = threading.Thread(target=self.run_scheduler, daemon=True)
        self.scheduler_thread.start()
        
    def create_shortcut_panel(self):
        """Membuat panel shortcut di sisi kiri aplikasi."""
        # Header
        header_label = ttk.Label(self.shortcut_frame, 
                               text="SHORTCUT PANEL",
                               font=("Helvetica", 14, "bold"),
                               foreground=self.style.primary_color)
        header_label.pack(pady=(0, 20))
        
        # Shortcut items
        shortcuts = [
            ("A", "Admin Login", "Masuk ke mode admin"),
            ("S", "Register Face", "Daftarkan wajah baru (Admin)"),
            ("E", "Export Report", "Ekspor laporan (Admin)"),
            ("X", "Exit Admin", "Keluar dari mode admin"),
            ("D", "Confirm Entry", "Konfirmasi kedatangan"),
            ("P", "Confirm Exit", "Konfirmasi kepulangan"),
            ("Q", "Quit", "Keluar dari aplikasi")
        ]
        
        for key, title, desc in shortcuts:
            # Container untuk setiap shortcut
            shortcut_container = ttk.Frame(self.shortcut_frame, style='TFrame')
            shortcut_container.pack(fill=tk.X, pady=5)
            
            # Key label dengan background
            key_label = ttk.Label(shortcut_container,
                                text=key,
                                font=("Helvetica", 12, "bold"),
                                foreground="white",
                                background=self.style.primary_color,
                                width=3)
            key_label.pack(side=tk.LEFT, padx=(0, 10))
            
            # Title dan description
            title_label = ttk.Label(shortcut_container,
                                  text=title,
                                  font=("Helvetica", 11, "bold"))
            title_label.pack(side=tk.LEFT, anchor=tk.W)
            
            desc_label = ttk.Label(shortcut_container,
                                 text=desc,
                                 font=("Helvetica", 9),
                                 foreground="gray")
            desc_label.pack(side=tk.LEFT, anchor=tk.W, padx=(5, 0))
        
        # Separator
        ttk.Separator(self.shortcut_frame, orient='horizontal').pack(fill=tk.X, pady=20)
        
        # Time info
        time_frame = ttk.Frame(self.shortcut_frame, style='TFrame')
        time_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(time_frame,
                 text="Waktu Presensi:",
                 font=("Helvetica", 11, "bold")).pack(anchor=tk.W)
        
        ttk.Label(time_frame,
                 text="Masuk: 05.00 - 09.00",
                 font=("Helvetica", 9)).pack(anchor=tk.W)
        
        ttk.Label(time_frame,
                 text=f"Pulang: {AFTERNOON_EXIT_START.strftime('%H:%M')} - {AFTERNOON_EXIT_END.strftime('%H:%M')}",
                 font=("Helvetica", 9)).pack(anchor=tk.W)
        
        # Admin mode indicator
        self.admin_indicator = ttk.Label(self.shortcut_frame,
                                       text="Mode: User",
                                       font=("Helvetica", 11, "bold"),
                                       foreground=self.style.primary_color)
        self.admin_indicator.pack(pady=20)
        
    def update_camera(self):
        """Update tampilan kamera dan overlay info wajah di frame kamera."""
        try:
            ret, frame = self.cap.read()
            if ret:
                frame = cv2.flip(frame, 1)
                self.current_frame = frame
                face_locations = face_recognition.face_locations(frame, model=FACE_DETECTION_MODEL)
                face_encodings = face_recognition.face_encodings(frame, face_locations)
                self.current_face_location = None
                self.current_face_encoding = None
                for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                    if self.current_face_location is None:
                        self.current_face_location = (top, right, bottom, left)
                        self.current_face_encoding = face_encoding
                    cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                    face_img = frame[top:bottom, left:right]
                    emotion = detect_emotion(face_img)
                    matches = face_recognition.compare_faces(
                        self.face_db.known_face_encodings,
                        face_encoding,
                        tolerance=FACE_MATCHING_TOLERANCE
                    )
                    if True in matches:
                        match_index = matches.index(True)
                        name = self.face_db.known_face_names[match_index]
                        role = self.face_db.known_face_roles[match_index]
                        level = self.face_db.known_face_levels[match_index]
                        y_offset = top - 10
                        if level:
                            cv2.putText(frame, f"Level: {level}", (left, y_offset - 40),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                        cv2.putText(frame, f"{name} ({role})", (left, y_offset - 20),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                        cv2.putText(frame, f"{emotion}", (left, y_offset - 60),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 200, 0), 2)
                    else:
                        y_offset = top - 10
                        cv2.putText(frame, "Unknown", (left, y_offset - 20),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                        cv2.putText(frame, f"{emotion}", (left, y_offset - 40),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 200, 0), 2)
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame_rgb)
                imgtk = ImageTk.PhotoImage(image=img)
                self.camera_label.imgtk = imgtk
                self.camera_label.configure(image=imgtk)
                if self.admin_mode:
                    self.admin_indicator.configure(text="Mode: Admin", foreground="red")
                else:
                    self.admin_indicator.configure(text="Mode: User", foreground=self.style.primary_color)
                self.root.update()
        except Exception as e:
            print(f"Error in update_camera: {e}")
        self.root.after(10, self.update_camera)
        
    def run_scheduler(self):
        """Menjalankan scheduler untuk tugas-tugas terjadwal."""
        while True:
            schedule.run_pending()
            time.sleep(1)
        
    def run(self):
        """Menjalankan aplikasi utama (mainloop Tkinter)."""
        try:
            self.root.mainloop()
        except Exception as e:
            print(f"Error in run: {e}")
        finally:
            self.cleanup()
            
    def handle_admin_login(self):
        """Menangani login admin."""
        if not self.admin_mode:
            self.show_admin_login()
            
    def handle_exit_admin(self):
        """Keluar dari mode admin."""
        if self.admin_mode:
            self.admin_mode = False
            print("Keluar dari mode admin")
            
    def handle_register_face(self):
        """Menangani pendaftaran wajah baru (hanya admin)."""
        if not self.admin_mode:
            messagebox.showwarning("Peringatan", "Anda harus login sebagai admin terlebih dahulu!")
            return
        if (
            not hasattr(self, 'current_face_location') or
            not hasattr(self, 'current_face_encoding') or
            self.current_face_location is None or
            self.current_face_encoding is None
        ):
            messagebox.showwarning("Peringatan", "Tidak ada wajah yang terdeteksi")
            return
        # Cek apakah wajah sudah terdaftar
        matches = face_recognition.compare_faces(
            self.face_db.known_face_encodings,
            self.current_face_encoding,
            tolerance=FACE_MATCHING_TOLERANCE
        )
        if any(matches):
            messagebox.showwarning("Peringatan", "Wajah anda sudah terdaftar")
            return
        self.register_new_face(self.current_frame, self.current_face_location, self.current_face_encoding)
            
    def handle_export_report(self):
        """Menangani ekspor laporan presensi dan screenshot."""
        if self.admin_mode:
            self.export_report()
            
    def show_admin_login(self):
        """Menampilkan window login admin."""
        login_window = AdminLoginWindow(self.root, self.face_db)
        self.root.wait_window(login_window.root)
        self.admin_mode = login_window.authenticated
        
    def register_new_face(self, frame, face_location, face_encoding):
        """Menampilkan window pendaftaran wajah baru."""
        top, right, bottom, left = face_location
        face_image = frame[top:bottom, left:right]
        emotion = detect_emotion(face_image)
        
        registration_window = RegistrationWindow(
            self.root, 
            face_image, 
            face_encoding,
            emotion
        )
        self.root.wait_window(registration_window.root)
        
        if registration_window.registration_complete:
            self.face_db.add_face(
                registration_window.name,
                registration_window.role,
                registration_window.level,
                face_image,
                face_encoding,
                emotion
            )
            self.face_db.load_known_faces()
            self.attendance_system.voice_engine.say(f"Face {registration_window.name}, has been registered")
            self.attendance_system.voice_engine.runAndWait()
            
    def export_report(self):
        """Ekspor laporan kehadiran dan screenshot ke Excel."""
        # Ekspor laporan kehadiran
        attendance_filename = self.face_db.export_to_excel()
        if attendance_filename:
            messagebox.showinfo("Export Success", f"Laporan kehadiran berhasil diekspor ke:\n{attendance_filename}")
        else:
            messagebox.showwarning("Export Failed", "Tidak ada data presensi untuk diekspor")
            
        # Ekspor laporan screenshot
        screenshot_filename = self.attendance_system.export_screenshot_report()
        if screenshot_filename:
            messagebox.showinfo("Export Success", f"Laporan screenshot berhasil diekspor ke:\n{screenshot_filename}")
        else:
            messagebox.showwarning("Export Failed", "Tidak ada data screenshot untuk diekspor")
    
    def cleanup(self):
        """Membersihkan resource sebelum aplikasi ditutup."""
        try:
            # Tutup kamera
            if hasattr(self, 'cap'):
                self.cap.release()
            
            # Tutup koneksi database
            if hasattr(self, 'face_db'):
                self.face_db.conn.close()
                self.face_db.admin_conn.close()
                self.face_db.attendance_conn.close()
            
            # Hentikan mesin suara
            if hasattr(self, 'attendance_system'):
                self.attendance_system.voice_engine.stop()
            
            # Tutup semua window
            cv2.destroyAllWindows()
            
        except Exception as e:
            print(f"Error in cleanup: {e}")
        finally:
            # Pastikan aplikasi ditutup
            self.root.quit()

    def display_instructions(self, frame):
        """(Tidak digunakan)"""
        pass

    def toggle_fullscreen(self):
        """Toggle window antara fullscreen dan normal."""
        self.is_fullscreen = not self.is_fullscreen
        self.root.attributes('-fullscreen', self.is_fullscreen)
        if self.is_fullscreen:
            self.fullscreen_btn.config(text="Restore")
        else:
            self.fullscreen_btn.config(text="Fullscreen")

    def update_time_label(self):
        """Update label waktu terkini di header aplikasi."""
        now = datetime.now().strftime('%A, %d %B %Y  %H:%M:%S')
        self.time_label.config(text=now)
        self.root.after(1000, self.update_time_label)

    def handle_key_press(self, event):
        """Menangani input keyboard untuk shortcut aplikasi."""
        key = event.char.lower()
        
        if key == 'a':
            self.handle_admin_login()
        elif key == 's':
            self.handle_register_face()
        elif key == 'e':
            self.handle_export_report()
        elif key == 'x':
            self.handle_exit_admin()
        elif key == 'd':
            self.handle_confirm_attendance()
        elif key == 'p':
            self.handle_confirm_exit()
        elif key == 'q':
            self.root.quit()

    def handle_confirm_attendance(self):
        """Menangani konfirmasi kehadiran manual (shortcut D)."""
        # Hanya bisa di mode user
        if self.admin_mode:
            messagebox.showwarning("Peringatan", "Konfirmasi kedatangan hanya dapat dilakukan di mode user!")
            return
        # Cek apakah ada wajah terdeteksi
        if (
            not hasattr(self, 'current_face_location') or
            not hasattr(self, 'current_face_encoding') or
            self.current_face_location is None or
            self.current_face_encoding is None
        ):
            messagebox.showwarning("Peringatan", "Tidak ada wajah terdeteksi!")
            return
        # Cek waktu kedatangan
        current_time = datetime.now().time()
        if not self.is_time_between(current_time, MORNING_ENTRY_START, MORNING_ENTRY_END):
            messagebox.showwarning(
                "Peringatan",
                f"Bukan waktu kehadiran!\nWaktu kedatangan: {MORNING_ENTRY_START.strftime('%H:%M')} - {MORNING_ENTRY_END.strftime('%H:%M')}"
            )
            return
        # Cek wajah terdaftar
        matches = face_recognition.compare_faces(
            self.face_db.known_face_encodings,
            self.current_face_encoding,
            tolerance=FACE_MATCHING_TOLERANCE
        )
        if not any(matches):
            messagebox.showwarning("Peringatan", "Wajah tidak terdaftar!")
            return
        best_match_index = np.argmin(face_recognition.face_distance(
            self.face_db.known_face_encodings,
            self.current_face_encoding
        ))
        if matches[best_match_index]:
            name = self.face_db.known_face_names[best_match_index]
            today = datetime.now().strftime('%Y-%m-%d')
            # Cek apakah sudah presensi masuk hari ini
            data_today = self.face_db.get_attendance_data(today)
            user_today = next((d for d in data_today if d['name'] == name and d['entry_time']), None)
            if user_today:
                entry_time = user_today.get('entry_time', '-')
                messagebox.showwarning("Peringatan", f"{name} telah tercatat masuk ke sekolah pada pukul {entry_time}. Silakan menunggu hingga esok hari untuk melakukan absensi kembali.")
                return
            emotion = detect_emotion(self.current_frame[
                self.current_face_location[0]:self.current_face_location[2],
                self.current_face_location[3]:self.current_face_location[1]
            ])
            self.show_simple_confirmation(
                self.current_frame,
                name,
                self.face_db.known_face_roles[best_match_index],
                self.face_db.known_face_levels[best_match_index],
                emotion,
                is_entry=True
            )

    def is_time_between(self, check_time, start_time, end_time):
        """Cek apakah waktu saat ini berada di antara waktu mulai dan selesai."""
        if start_time <= end_time:
            return start_time <= check_time <= end_time
        else:  # Waktu melewati tengah malam
            return start_time <= check_time or check_time <= end_time

    def show_simple_confirmation(self, frame, name, role, level, emotion, is_entry=True, durasi=None):
        """Menampilkan window konfirmasi sederhana untuk kehadiran/kepulangan."""
        today = datetime.now().strftime('%Y-%m-%d')
        key = f"{name}_{today}"
        last_time = self.last_manual_confirmation.get(key, datetime.min)
        time_diff = datetime.now() - last_time
        if time_diff.total_seconds() < 300:
            messagebox.showwarning(
                "Peringatan",
                f"Harap tunggu {int(300 - time_diff.total_seconds())} detik lagi sebelum konfirmasi berikutnya."
            )
            return
        confirmation_window = SimpleConfirmationWindow(self.main_frame, name, is_entry, durasi)
        self.main_frame.wait_window(confirmation_window.root)
        if confirmation_window.confirmed:
            self.face_db.record_attendance(
                name, role, level, 
                is_entry=is_entry, 
                emotion=emotion,
                keterangan="Manual Entry"
            )
            self.last_manual_confirmation[key] = datetime.now()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            status = "ENTRY" if is_entry else "EXIT"
            filename = f"attendance_screenshots/MANUAL_{status}_{name}_{timestamp}.jpg"
            cv2.imwrite(filename, frame)
            if is_entry:
                self.attendance_system.voice_engine.say(f"Welcome {name}, Your current emotion is {emotion}. Good luck with your studies.")
                self.attendance_system.voice_engine.runAndWait()
            else:
                # Suara pulang
                self.attendance_system.voice_engine.say(f"Good bye {name}")
                self.attendance_system.voice_engine.say(f"Your emotion {emotion}")
                self.attendance_system.voice_engine.say("Be careful on the way")
                self.attendance_system.voice_engine.runAndWait()
            messagebox.showinfo(
                "Sukses",
                f"{'Kehadiran' if is_entry else 'Kepulangan'} {name} berhasil dicatat pada {datetime.now().strftime('%H:%M:%S')}"
            )

    def handle_confirm_exit(self):
        """Menangani konfirmasi kepulangan manual (shortcut P)."""
        # Hanya bisa di mode user
        if self.admin_mode:
            messagebox.showwarning("Peringatan", "Konfirmasi kepulangan hanya dapat dilakukan di mode user!")
            return
        # Cek apakah ada wajah terdeteksi
        if (
            not hasattr(self, 'current_face_location') or
            not hasattr(self, 'current_face_encoding') or
            self.current_face_location is None or
            self.current_face_encoding is None
        ):
            messagebox.showwarning("Peringatan", "Tidak ada wajah terdeteksi!")
            return
        # Cek waktu kepulangan
        current_time = datetime.now().time()
        if not self.is_time_between(current_time, AFTERNOON_EXIT_START, AFTERNOON_EXIT_END):
            messagebox.showwarning(
                "Peringatan", 
                f"Bukan waktu kepulangan!\nWaktu kepulangan: {AFTERNOON_EXIT_START.strftime('%H:%M')} - {AFTERNOON_EXIT_END.strftime('%H:%M')}"
            )
            return
        # Cek wajah terdaftar
        matches = face_recognition.compare_faces(
            self.face_db.known_face_encodings,
            self.current_face_encoding,
            tolerance=FACE_MATCHING_TOLERANCE
        )
        if not any(matches):
            messagebox.showwarning("Peringatan", "Wajah tidak terdaftar!")
            return
        best_match_index = np.argmin(face_recognition.face_distance(
            self.face_db.known_face_encodings,
            self.current_face_encoding
        ))
        if matches[best_match_index]:
            name = self.face_db.known_face_names[best_match_index]
            today = datetime.now().strftime('%Y-%m-%d')
            now = datetime.now()
            # Ambil data presensi hari ini
            data_today = self.face_db.get_attendance_data(today)
            user_today = next((d for d in data_today if d['name'] == name and d['exit_time']), None)
            if user_today:
                # Sudah presensi pulang hari ini, cek apakah sudah boleh presensi lagi
                last_exit = datetime.strptime(f"{today} {user_today['exit_time']}", "%Y-%m-%d %H:%M:%S")
                next_allowed = (last_exit + timedelta(days=1)).replace(hour=13, minute=0, second=0)
                if now < next_allowed:
                    sisa = next_allowed - now
                    jam, sisa = divmod(sisa.seconds, 3600)
                    menit, detik = divmod(sisa, 60)
                    messagebox.showwarning("Peringatan", f"Anda baru bisa presensi pulang lagi setelah besok jam 13.00.\nSisa waktu: {jam:02d}:{menit:02d}:{detik:02d}")
                    return
            emotion = detect_emotion(self.current_frame[
                self.current_face_location[0]:self.current_face_location[2],
                self.current_face_location[3]:self.current_face_location[1]
            ])
            # Ambil durasi dari record hari ini jika ada
            durasi = user_today['exit_time'] and user_today.get('status') == 'Pulang' and user_today.get('duration') or None
            self.show_simple_confirmation(
                self.current_frame,
                name,
                self.face_db.known_face_roles[best_match_index],
                self.face_db.known_face_levels[best_match_index],
                emotion,
                is_entry=False,
                durasi=durasi
            )

# ========== REGISTRATION WINDOW ==========
class RegistrationWindow:
    """Window untuk pendaftaran wajah baru (input nama, role, level)."""
    def __init__(self, parent, face_image, face_encoding, emotion):
        self.parent = parent
        self.face_image = face_image
        self.face_encoding = face_encoding
        self.emotion = emotion
        self.style = AppStyle()
        self.registration_complete = False
        self.name = ""
        self.role = ""
        self.level = ""
        
        self.root = tk.Toplevel(parent)
        self.root.title("Register New Face")
        self.root.geometry("500x650")  # Diperbesar untuk menambahkan level
        self.root.resizable(False, False)
        self.style.apply_style(self.root)
        
        self.create_widgets()
        
    def create_widgets(self):
        """Membuat dan menata widget pada window registrasi wajah baru."""
        # Header
        header_frame = ttk.Frame(self.root, style='TFrame')
        header_frame.pack(fill=tk.X, padx=20, pady=20)
        
        ttk.Label(header_frame, text="NEW REGISTRATION", 
                 font=self.style.title_font).pack()
        
        # Preview wajah
        preview_frame = ttk.Frame(self.root, style='TFrame')
        preview_frame.pack(pady=10)
        
        face_img = Image.fromarray(cv2.cvtColor(self.face_image, cv2.COLOR_BGR2RGB))
        face_img = face_img.resize((200, 200), Image.LANCZOS)
        self.face_preview = ImageTk.PhotoImage(face_img)
        
        ttk.Label(preview_frame, image=self.face_preview).pack()
        ttk.Label(preview_frame, text=f"Emotion: {self.emotion}", 
                 font=self.style.subtitle_font).pack(pady=5)
        
        # Form
        form_frame = ttk.Frame(self.root, style='TFrame')
        form_frame.pack(fill=tk.X, padx=40, pady=10)
        
        # Name
        ttk.Label(form_frame, text="Full Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.name_entry = ttk.Entry(form_frame, font=self.style.normal_font)
        self.name_entry.grid(row=0, column=1, sticky=tk.EW, pady=5)
        
        # Role
        ttk.Label(form_frame, text="Role:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.role_combobox = ttk.Combobox(
            form_frame, 
            values=ROLE_OPTIONS, 
            state="readonly",
            font=self.style.normal_font
        )
        self.role_combobox.current(0)
        self.role_combobox.grid(row=1, column=1, sticky=tk.EW, pady=5)
        
        # Level (hanya untuk siswa)
        ttk.Label(form_frame, text="Level (for students):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.level_combobox = ttk.Combobox(
            form_frame, 
            values=LEVEL_OPTIONS, 
            state="readonly",
            font=self.style.normal_font
        )
        self.level_combobox.current(0)
        self.level_combobox.grid(row=2, column=1, sticky=tk.EW, pady=5)
        
        # Error message
        self.error_label = ttk.Label(form_frame, text="", style='Error.TLabel')
        self.error_label.grid(row=3, column=0, columnspan=2, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(self.root, style='TFrame')
        button_frame.pack(fill=tk.X, padx=40, pady=20)
        
        ttk.Button(button_frame, text="REGISTER", command=self.register, 
                  style='Accent.TButton').pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="CANCEL", command=self.cancel).pack(fill=tk.X)
        
        self.root.grab_set()
    
    def register(self):
        """Validasi dan simpan data pendaftaran wajah baru."""
        self.name = self.name_entry.get().strip()
        self.role = self.role_combobox.get()
        self.level = self.level_combobox.get() if self.role == "Siswa" else ""
        
        if not self.name:
            self.error_label.config(text="Full name must be filled!")
            return
            
        self.registration_complete = True
        self.root.destroy()
    
    def cancel(self):
        """Tutup window registrasi tanpa menyimpan."""
        self.root.destroy()

# ========== CONFIRMATION WINDOW ==========
class ConfirmationWindow:
    """Window konfirmasi kehadiran/kepulangan dengan detail info."""
    def __init__(self, parent, name, role, level, emotion, is_entry=True):
        self.parent = parent
        self.name = name
        self.role = role
        self.level = level
        self.emotion = emotion
        self.is_entry = is_entry
        self.style = AppStyle()
        self.confirmed = False
        self.keterangan = None
        
        self.root = tk.Toplevel(parent)
        self.root.title("Konfirmasi Presensi")
        self.root.geometry("400x500")
        self.root.resizable(False, False)
        self.style.apply_style(self.root)
        
        self.create_widgets()
        
    def create_widgets(self):
        """Membuat dan menata widget pada window konfirmasi detail."""
        # Header
        header_frame = ttk.Frame(self.root, style='TFrame')
        header_frame.pack(fill=tk.X, padx=20, pady=20)
        
        title = "KONFIRMASI KEDATANGAN" if self.is_entry else "KONFIRMASI KEPULANGAN"
        ttk.Label(header_frame, text=title, 
                 font=self.style.title_font).pack()
        
        # Info frame
        info_frame = ttk.Frame(self.root, style='TFrame')
        info_frame.pack(fill=tk.X, padx=40, pady=20)
        
        # Tampilkan informasi
        ttk.Label(info_frame, text=f"Nama: {self.name}", 
                 font=self.style.subtitle_font).pack(anchor=tk.W, pady=5)
        ttk.Label(info_frame, text=f"Peran: {self.role}", 
                 font=self.style.subtitle_font).pack(anchor=tk.W, pady=5)
        if self.level:
            ttk.Label(info_frame, text=f"Level: {self.level}", 
                     font=self.style.subtitle_font).pack(anchor=tk.W, pady=5)
        ttk.Label(info_frame, text=f"Emosi: {self.emotion}", 
                 font=self.style.subtitle_font).pack(anchor=tk.W, pady=5)
        
        # Keterangan (untuk keterlambatan)
        if self.is_entry:
            ttk.Label(info_frame, text="Keterangan (opsional):", 
                     font=self.style.subtitle_font).pack(anchor=tk.W, pady=5)
            self.keterangan_entry = ttk.Entry(info_frame, font=self.style.normal_font)
            self.keterangan_entry.pack(fill=tk.X, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(self.root, style='TFrame')
        button_frame.pack(fill=tk.X, padx=40, pady=20)
        
        ttk.Button(button_frame, text="KONFIRMASI", command=self.confirm, 
                  style='Accent.TButton').pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="BATAL", command=self.cancel).pack(fill=tk.X)
        
        self.root.grab_set()
    
    def confirm(self):
        """Konfirmasi kehadiran/kepulangan."""
        self.keterangan = self.keterangan_entry.get().strip() if self.is_entry else None
        self.confirmed = True
        self.root.destroy()
    
    def cancel(self):
        """Tutup window konfirmasi tanpa menyimpan."""
        self.root.destroy()

# ========== SIMPLE CONFIRMATION WINDOW ==========
class SimpleConfirmationWindow:
    """Window konfirmasi sederhana (hanya nama dan waktu)."""
    def __init__(self, parent, name, is_entry=True, durasi=None):
        self.parent = parent
        self.name = name
        self.is_entry = is_entry
        self.durasi = durasi
        self.style = AppStyle()
        self.confirmed = False
        
        self.root = tk.Toplevel(parent)
        self.root.title("Konfirmasi")
        self.root.geometry("400x300")
        self.root.resizable(False, False)
        self.style.apply_style(self.root)
        
        # Buat window muncul di tengah layar
        self.root.transient(parent)
        self.root.grab_set()
        
        # Posisikan window di tengah
        x = parent.winfo_x() + (parent.winfo_width() - 400) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 300) // 2
        self.root.geometry(f"+{x}+{y}")
        
        self.create_widgets()
        
    def create_widgets(self):
        """Membuat dan menata widget pada window konfirmasi sederhana."""
        # Main container dengan padding
        main_container = ttk.Frame(self.root, style='TFrame')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        # Header dengan background
        header_frame = ttk.Frame(main_container, style='TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 20))
        title = "KONFIRMASI KEDATANGAN" if self.is_entry else "KONFIRMASI KEPULANGAN"
        ttk.Label(header_frame, text=title, 
                 font=("Helvetica", 16, "bold"),
                 foreground="#4a6fa5").pack()
        # Info frame dengan border
        info_frame = ttk.Frame(main_container, style='TFrame')
        info_frame.pack(fill=tk.X, pady=10)
        # Tampilkan pertanyaan dengan style yang lebih baik
        ttk.Label(info_frame, 
                 text="Apakah benar ini adalah:",
                 font=("Helvetica", 12)).pack(pady=(0, 5))
        ttk.Label(info_frame, 
                 text=self.name,
                 font=("Helvetica", 14, "bold"),
                 foreground="#166088").pack(pady=(0, 10))
        # Waktu saat ini
        current_time = datetime.now().strftime('%H:%M:%S')
        ttk.Label(info_frame,
                 text=f"Waktu: {current_time}",
                 font=("Helvetica", 12)).pack(pady=(0, 10))
        # Tampilkan durasi jika kepulangan dan ada durasi
        if not self.is_entry and self.durasi:
            ttk.Label(info_frame,
                     text=f"Durasi di sekolah: {self.durasi}",
                     font=("Helvetica", 12, "italic"),
                     foreground="#388e3c").pack(pady=(0, 10))
        # Button frame dengan style yang lebih baik
        button_frame = ttk.Frame(main_container, style='TFrame')
        button_frame.pack(fill=tk.X, pady=(20, 0))
        # Tombol YA dengan style accent
        ttk.Button(button_frame, 
                  text="YA, BENAR",
                  command=self.confirm,
                  style='Accent.TButton').pack(fill=tk.X, pady=(0, 10))
        # Tombol TIDAK dengan style yang berbeda
        ttk.Button(button_frame,
                  text="TIDAK, BATAL",
                  command=self.cancel).pack(fill=tk.X)
        # Bind Enter key untuk konfirmasi
        self.root.bind('<Return>', lambda e: self.confirm())
        # Bind Escape key untuk batal
        self.root.bind('<Escape>', lambda e: self.cancel())
        # Fokus ke window
        self.root.focus_set()
    
    def confirm(self):
        """Konfirmasi kehadiran/kepulangan sederhana."""
        self.confirmed = True
        self.root.destroy()
    
    def cancel(self):
        """Tutup window konfirmasi sederhana tanpa menyimpan."""
        self.root.destroy()

# ========== FUNGSI UTILITAS ==========
def detect_emotion(face_image):
    """Mendeteksi ekspresi wajah menggunakan DeepFace."""
    try:
        rgb_image = cv2.cvtColor(face_image, cv2.COLOR_BGR2RGB)
        result = DeepFace.analyze(rgb_image, actions=['emotion'], enforce_detection=False)
        
        if isinstance(result, list):
            result = result[0]
            
        return result['dominant_emotion'].capitalize()
    except Exception as e:
        print(f"Error deteksi ekspresi: {e}")
        return "Unknown"

if __name__ == "__main__":
    app = FaceRecognitionApp()
    app.run()